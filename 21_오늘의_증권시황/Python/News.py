
import os
import requests
import pandas as pd
from bs4 import BeautifulSoup as BSoup
from openpyxl import load_workbook


def get_news_dataframe(day):
    url_base = 'https://finance.naver.com'
    url = url_base + f'/news/mainnews.naver?date={day}'
    
    data = []
    page_num = 1
    while True:
        html = requests.get(f'{url}&page={page_num}').text
        soup = BSoup(html, 'html.parser')

        for news in soup.find_all('li', 'block1'):
            subject = news.find('dd', 'articleSubject')
            title = subject.text.strip()
            link = url_base + subject.a['href']
            summary = news.find('dd', 'articleSummary')
            content = summary.contents[0].strip()
            press = summary.find('span', 'press').text
            date = summary.find('span', 'wdate').text
            data.append([f'=HYPERLINK("{link}", "▶")', title, content, press, date])
            # print(link, title, content, press, date, sep='\n')

        page_num += 1
        # 다음 페이지가 없으면 종료
        if soup.select_one('.pgRR') is None:
            break
    return pd.DataFrame(data, columns=['링크', '제목', '내용', '언론사', '날짜'])

def insert_news_file(filePath, sheetName, startRow, startCol, day):
    try:
        df = get_news_dataframe(day)
        result = f'{sheetName}를 가져왔습니다. ({len(df)}행)'
    except Exception as e:
        result = f'에러 발생: {e}'
        result = f"{result}, \n {filePath} \n {sheetName} ({startRow}, {startCol}) {day}"
        return result
    finally:
        print(result)

    filePath = convert_path(filePath)
    try:
        wb = load_workbook(filePath)
        ws = wb[sheetName]
        # 기존 셀 삭제
        # for row in ws.iter_rows():
        #     for cell in row:
        #         cell.value = None
        # 제목 삽입
        for idx, colTitle in enumerate(df.columns, start=startCol):
            ws.cell(row=startRow, column=idx, value=colTitle)
        # 값 삽입
        for r, row in enumerate(df.values, start=startRow+1):
            for c, value in enumerate(row, start=startCol):
                ws.cell(row=r, column=c, value=value)
        wb.save(filePath)
        result = f'{sheetName}가 정상적으로 삽입되었습니다. ({day})'
    except Exception as e:
        result = f'에러 발생: {e}'
    print(result)
    return result


def convert_path(*paths):
    '''
    입력된 경로 전체의 폴더 구분자를 '/'으로 변환하여 반환
    paths <*str> : 경로 문자열을 가변매개변수로 입력
    반환값 <tuple(str) | str | None> : 입력이 다수이면 튜플로, 하나이면 문자열로, 없으면 None을 반환
    '''
    result = tuple(path.replace(os.sep, '/') for path in paths)
    return result if len(paths) > 1 else result[0] if len(paths) == 1 else None 



# 연습예제
if __name__ == '__main__':
    fileFolder = "C:\\Users\\Dexter\\Source\\ALPACO8\\RPA_Example\\21_오늘의_증권시황_스크린샷\\Data\\Output"
    filePath = os.path.join(fileFolder, "Sample.xlsx")
    # insert_news_file(filePath, '주요뉴스', 5, 3, '2024-12-18')